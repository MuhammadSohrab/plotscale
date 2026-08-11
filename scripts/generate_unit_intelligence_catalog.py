"""Generate PlotScale's research-safe country and unit-intelligence indexes.

This script reads research workbooks only. It never copies legacy UI, state, or
application architecture, and it never promotes a researched factor into a
runtime conversion factor.
"""

from __future__ import annotations

import json
import re
import sys
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_RESEARCH_ROOT = PROJECT_ROOT.parent / "_legacy_reference" / "unit_research"
OUTPUT_ROOT = PROJECT_ROOT / "src" / "data" / "generated"
REPORT_ROOT = PROJECT_ROOT / "docs" / "unit-intelligence"

LAND_TERMS = {
    "acre", "area", "bigha", "biswa", "cent", "chain", "chatak", "dhur",
    "dunam", "field", "gunta", "guntha", "hectare", "jareeb", "jerib",
    "kanal", "karam", "katha", "katt", "land", "laggi", "lessa", "maal",
    "manzana", "marla", "morgen", "mu", "ngan", "paisa", "perch", "ping",
    "plot", "rai", "rod", "ropani", "square", "stremma", "survey", "tarea",
    "tarang", "wah", "yard",
}
NON_LAND_TERMS = {
    "astronomical", "light-year", "parsec", "nautical", "pixel", "typographic",
    "electron", "atomic", "wavelength",
}


def clean(value):
    return "" if value is None else str(value).strip()


def slug(value):
    return re.sub(r"[^a-z0-9]+", "_", clean(value).lower()).strip("_")


def locate_header(rows):
    for index, row in enumerate(rows[:12]):
        normalized = {slug(value) for value in row if value is not None}
        if len(normalized) >= 2 and (
            {"unit_id", "unit_name"} & normalized
            or {"country_name", "iso2"} <= normalized
            or {"conversion_key", "display_name"} <= normalized
            or "relationship_id" in normalized
        ):
            return index
    return None


def rows_as_dicts(path, sheet_name):
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[sheet_name]
    rows = list(sheet.iter_rows(values_only=True))
    header_index = locate_header(rows)
    if header_index is None:
        return []
    headers = [slug(value) or f"column_{index}" for index, value in enumerate(rows[header_index])]
    output = []
    for row_number, row in enumerate(rows[header_index + 1 :], start=header_index + 2):
        record = {headers[index]: row[index] if index < len(row) else None for index in range(len(headers))}
        if any(clean(value) for value in record.values()):
            record["_row"] = row_number
            output.append(record)
    return output


def first(record, *keys):
    for key in keys:
        value = record.get(key)
        if clean(value):
            return clean(value)
    return ""


def infer_dimension(record):
    explicit = first(
        record,
        "dimension",
        "unit_dimension",
        "measure_dimension",
        "quantity_kind",
        "measurement_type",
    ).lower()
    if explicit:
        if any(term in explicit for term in ("area", "square", "surface")):
            return "area"
        if any(term in explicit for term in ("length", "distance", "linear", "tool")):
            return "length"
    unit_name = first(
        record,
        "unit_name",
        "display_name",
        "canonical_name",
        "unit",
        "unit_concept",
        "conversion_key",
    ).lower()
    if re.search(r"\b(square|sq\.?|area)\b", unit_name):
        return "area"
    if any(term in unit_name for term in (
        "karam", "laggi", "measuring rope", "linear chain", "length jareeb",
    )):
        return "length"
    text = " ".join(clean(value).lower() for value in record.values())
    if any(term in text for term in ("area", "sq ", "square", "m2", "m²", "acre", "hectare")):
        return "area"
    if any(term in text for term in ("length", "metre", "meter", "foot", "yard", "chain", "rod")):
        return "length"
    return "unknown"


def classify_role(record):
    text = " ".join(clean(value).lower() for value in record.values())
    if any(term in text for term in NON_LAND_TERMS):
        return "rejected_non_land"
    if "non-geometric" in text or "taxation" in text or "seed" in text or "productivity" in text:
        return "non_geometric_reference_only"
    if "ambiguous" in text or "polysem" in text or "conflict" in text:
        return "ambiguous_reference_only"
    if "histor" in text or "obsolete" in text or "ancient" in text:
        return "historical_geometric_candidate"
    if any(key in record for key in ("parent_unit", "hierarchy_parent", "from_unit_id", "to_unit_id")):
        if first(record, "hierarchy_parent", "parent_unit", "from_unit_id", "to_unit_id"):
            return "relationship_candidate"
    if "tool" in text or "rope" in text or "laggi" in text or "karam" in text or "chain" in text:
        return "measuring_tool_candidate"
    if first(record, "aliases", "aliases_en", "native_or_alias"):
        return "alias_candidate"
    dimension = infer_dimension(record)
    if dimension in {"length", "area"}:
        return "family_member_candidate"
    return "standalone_local_unit_candidate"


def is_land_relevant(record):
    text = " ".join(clean(value).lower() for value in record.values())
    explicit = first(record, "land_relevance")
    if explicit:
        return explicit.lower() not in {"no", "false", "0", "not land-relevant"}
    return any(term in text for term in LAND_TERMS)


def build_country_directory(worldwide_path):
    records = rows_as_dicts(worldwide_path, "Country_Coverage")
    countries = []
    for item in records:
        iso2 = first(item, "iso2").upper()
        if len(iso2) != 2:
            continue
        countries.append({
            "code": iso2,
            "iso3": first(item, "iso3").upper(),
            "numericCode": first(item, "numeric_code").zfill(3),
            "name": first(item, "country_name"),
            "continent": first(item, "continent"),
            "coverageStatus": first(item, "coverage_status"),
            "researchRecordCount": int(item.get("wikidata_or_variant_rows") or 0),
            "source": "Worldwide_Land_Measurement_Database.xlsx#Country_Coverage",
        })
    return sorted(countries, key=lambda item: item["name"])


def source_specs(root):
    return [
        (root / "Worldwide_Land_Measurement_Database.xlsx", [
            "Converter_Index", "Sizes_Snippets", "India_Wikipedia", "Master_Units",
            "WD_Conversions", "Unit_Relations", "QUDT_Standard",
        ]),
        (root / "land_unit_presets.xlsx", ["units", "relationships"]),
        (root / "PlotScale_Global_Units_Research.xlsx", [
            "Unit_Master", "SizesCom_Extracts", "Country_Locality", "Hierarchies",
            "Length_Units", "Area_Units", "Europe_Historical", "SouthAsia_Detail",
        ]),
        (root / "PlotScale_Land_Units_Research.xlsx", [
            "Common_Units", "HindiWiki_Data", "SizesCom_Data", "Regional_Area_Units",
            "State_Unit_Map_NGDRS",
        ]),
        (root / "PlotScale_Units_Family_Structure_Review.xlsx", [
            "Fixed_Units", "MultiSize_Families", "MultiSize_Ratios", "Measuring_Tools",
            "Location_Estimates", "Research_Advanced",
        ]),
    ]


def build_classification(root, countries):
    country_by_name = {item["name"].lower(): item["code"] for item in countries}
    country_by_code = {item["code"]: item["code"] for item in countries}
    country_aliases = {
        "usa": "US", "u.s.": "US", "united states": "US",
        "uk": "GB", "england": "GB", "scotland": "GB", "britain": "GB",
        "south korea": "KR", "korea": "KR", "north korea": "KP",
        "russia": "RU", "vietnam": "VN", "taiwan": "TW",
        "palestine": "PS", "macau": "MO", "macao": "MO",
        "syria": "SY", "turkey": "TR", "moldavia": "MD",
        "bolivia": "BO", "venezuela": "VE", "laos": "LA",
        "iran": "IR", "tanzania": "TZ", "brunei": "BN",
    }

    def resolve_country_codes(label):
        normalized = clean(label).lower()
        matches = set()
        direct = country_by_code.get(clean(label).upper()) or country_by_name.get(normalized)
        if direct:
            matches.add(direct)
        for name, code in country_by_name.items():
            if re.search(rf"(?<![a-z]){re.escape(name)}(?![a-z])", normalized):
                matches.add(code)
        for name, code in country_aliases.items():
            if re.search(rf"(?<![a-z]){re.escape(name)}(?![a-z])", normalized):
                matches.add(code)
        return sorted(matches)

    classification = []
    rejected = []
    for workbook_path, sheets in source_specs(root):
        if not workbook_path.exists():
            continue
        workbook = load_workbook(workbook_path, read_only=True, data_only=True)
        available = set(workbook.sheetnames)
        workbook.close()
        for sheet_name in sheets:
            if sheet_name not in available:
                continue
            for record in rows_as_dicts(workbook_path, sheet_name):
                if not is_land_relevant(record):
                    continue
                country = first(
                    record, "country_code", "country", "country_or_region",
                    "countries_stated", "region", "state_ut",
                )
                country_codes = resolve_country_codes(country)
                country_code = country_codes[0] if len(country_codes) == 1 else None
                unit_name = first(
                    record, "unit_name", "unit_name_en", "display_name",
                    "relationship_unit", "unit_id_name", "unitfamily",
                ) or first(record, "unit_id", "conversion_key", "relationship_id")
                role = classify_role(record)
                dimension = infer_dimension(record)
                if role == "measuring_tool_candidate":
                    dimension = "length"
                entry = {
                    "sourceRecordId": f"{slug(workbook_path.stem)}:{slug(sheet_name)}:{record['_row']}",
                    "sourceFile": workbook_path.name,
                    "sourceSheet": sheet_name,
                    "sourceRow": record["_row"],
                    "countryCode": country_code,
                    "countryCodes": country_codes,
                    "locationLabel": country or first(record, "locality", "region_label", "admin1", "district"),
                    "unitConcept": unit_name,
                    "dimension": dimension,
                    "role": role,
                    "suggestionPackId": (
                        f"research_{country_code.lower()}_candidates"
                        if country_code else
                        f"research_multi_region_{slug(country)}" if country_codes else
                        "research_unscoped_reference_candidates"
                    ),
                    "runtimeBehavior": "reference_only" if role.endswith("reference_only") else "suggestion_only",
                    "reason": "Research assertion retained for user-confirmed setup; no factor is activated.",
                    "hasConflict": "conflict" in " ".join(clean(value).lower() for value in record.values()),
                    "sourceLocator": f"{workbook_path.name}#{sheet_name}!{record['_row']}",
                    "aliases": [
                        item.strip()
                        for item in re.split(
                            r"[,;/|]",
                            first(record, "aliases", "aliases_en", "native_or_alias"),
                        )
                        if item.strip()
                    ],
                    "proposedRelationship": {
                        "parent": first(record, "hierarchy_parent", "parent_unit", "from_unit_id"),
                        "child": first(record, "hierarchy_child", "child_unit", "to_unit_id"),
                        "multiplier": first(
                            record,
                            "multiplier",
                            "units_per_parent",
                            "ratio",
                            "conversion_multiplier",
                        ),
                    } if first(
                        record,
                        "hierarchy_parent",
                        "parent_unit",
                        "from_unit_id",
                        "to_unit_id",
                    ) else None,
                }
                if role == "rejected_non_land":
                    rejected.append(entry)
                else:
                    classification.append(entry)
    return classification, rejected


def write_json(path, value):
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(value, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def main():
    root = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_RESEARCH_ROOT
    worldwide = root / "Worldwide_Land_Measurement_Database.xlsx"
    if not worldwide.exists():
        raise SystemExit(f"Research workbook not found: {worldwide}")

    countries = build_country_directory(worldwide)
    classified, rejected = build_classification(root, countries)
    coverage_counts = Counter(
        code
        for item in classified
        for code in item.get("countryCodes", [])
    )
    coverage = [{
        "countryCode": country["code"],
        "recordCount": coverage_counts.get(country["code"], 0),
        "hasResearchSuggestions": coverage_counts.get(country["code"], 0) > 0,
    } for country in countries]
    role_counts = Counter(item["role"] for item in classified)
    grouped_packs = {}
    for item in classified:
        pack_id = item["suggestionPackId"]
        pack = grouped_packs.setdefault(pack_id, {
            "id": pack_id,
            "version": "1.0.0",
            "tier": "suggested",
            "runtimeBehavior": "user_setup_only",
            "countryCodes": set(),
            "locationLabels": set(),
            "candidates": [],
        })
        pack["countryCodes"].update(item.get("countryCodes", []))
        if item.get("locationLabel"):
            pack["locationLabels"].add(item["locationLabel"])
        pack["candidates"].append({
            "sourceRecordId": item["sourceRecordId"],
            "unitConcept": item["unitConcept"],
            "dimension": item["dimension"],
            "role": item["role"],
            "locationLabel": item["locationLabel"],
            "hasConflict": item["hasConflict"],
            "sourceLocator": item["sourceLocator"],
            "runtimeBehavior": item["runtimeBehavior"],
            "aliases": item["aliases"],
            "proposedRelationship": item["proposedRelationship"],
        })
    suggestion_catalog = []
    for pack in grouped_packs.values():
        countries_for_pack = sorted(pack["countryCodes"])
        pack["countryCodes"] = countries_for_pack
        pack["locationLabels"] = sorted(pack["locationLabels"])
        pack["name"] = (
            f"Additional researched unit candidates — {countries_for_pack[0]}"
            if len(countries_for_pack) == 1
            else "Additional researched multi-region unit candidates"
            if countries_for_pack
            else "Unscoped historical and ambiguous unit references"
        )
        pack["candidateCount"] = len(pack["candidates"])
        pack["roleCounts"] = dict(Counter(item["role"] for item in pack["candidates"]))
        pack["hasConflicts"] = any(item["hasConflict"] for item in pack["candidates"])
        pack["warning"] = (
            "Names and possible relationships only. Research factors are not "
            "available to the conversion engine."
        )
        suggestion_catalog.append(pack)
    suggestion_catalog.sort(key=lambda item: item["id"])
    suggestion_index = [{
        key: value
        for key, value in pack.items()
        if key != "candidates"
    } for pack in suggestion_catalog]

    write_json(OUTPUT_ROOT / "countryDirectory.json", countries)
    write_json(OUTPUT_ROOT / "researchClassification.json", classified)
    write_json(OUTPUT_ROOT / "researchRejections.json", rejected)
    write_json(OUTPUT_ROOT / "suggestionCoverage.json", coverage)
    write_json(OUTPUT_ROOT / "researchSuggestionCatalog.json", suggestion_catalog)
    write_json(OUTPUT_ROOT / "researchSuggestionCatalogIndex.json", suggestion_index)

    REPORT_ROOT.mkdir(parents=True, exist_ok=True)
    report = [
        "# PlotScale Unit Research Classification",
        "",
        "Generated audit metadata only. No research factor is enabled for conversion.",
        "",
        f"- Countries/territories selectable: {len(countries)}",
        f"- Land-relevant classified assertions: {len(classified)}",
        f"- Rejected non-land assertions: {len(rejected)}",
        f"- Countries with mapped research assertions: {sum(1 for item in coverage if item['hasResearchSuggestions'])}",
        f"- Generated suggestion/reference packs: {len(suggestion_catalog)}",
        "",
        "## Roles",
        "",
        *[f"- `{role}`: {count}" for role, count in sorted(role_counts.items())],
        "",
        "See generated JSON files for record-level source locators and classifications.",
    ]
    (REPORT_ROOT / "classification-report.md").write_text("\n".join(report) + "\n", encoding="utf-8")
    print("\n".join(report))


if __name__ == "__main__":
    main()
